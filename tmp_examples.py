from openpyxl import load_workbook
from pathlib import Path

excel_file = Path(r"c:/1027/data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v3.4.xlsx")
wb = load_workbook(excel_file, data_only=False)
ws = wb.active
orange_examples = []
yellow_examples = []
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        if fill.start_color is not None:
            rgb = str(fill.start_color.rgb or "").upper()
        else:
            rgb = ""
        if not rgb:
            continue
        if "FFC000" in rgb and len(orange_examples) < 3:
            orange_examples.append((row, col, rgb))
        elif "FFFF00" in rgb and len(yellow_examples) < 3:
            yellow_examples.append((row, col, rgb))
    if len(orange_examples) >= 3 and len(yellow_examples) >= 3:
        break
print("orange_examples", orange_examples)
print("yellow_examples", yellow_examples)
wb.close()